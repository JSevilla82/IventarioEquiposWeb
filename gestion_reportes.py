# gestion_reportes.py
import webbrowser
from datetime import datetime
import tempfile

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from database import db_manager, Equipo, registrar_movimiento_sistema
from gestion_acceso import requiere_permiso

# --- MENÚ PRINCIPAL DE VISUALIZACIÓN ---
@requiere_permiso("ver_inventario")
def menu_ver_inventario(usuario: str):
    """Menú principal para la visualización (adaptado para web)."""
    print("\n--- Módulo de Visualización de Inventario ---")
    opciones = [
        "1. Generar Reportes de Inventario en Excel",
        "2. Ver últimos 20 movimientos",
        "3. Ver Inventario Actual en Consola",
        "4. Volver al menú principal"
    ]
    print("\n".join(opciones))
    # La lógica para manejar la opción del usuario se gestionará en app.py

# --- SUBMENÚS Y FUNCIONES DE VISUALIZACIÓN ---

def menu_ver_ultimos_movimientos(usuario: str):
    """Muestra una tabla con los últimos 20 movimientos de inventario del usuario."""
    movimientos = db_manager.get_last_movimientos_by_user(usuario, limit=20)
    
    print("\n--- Tus Últimos 20 Movimientos ---")

    if not movimientos:
        print("No has registrado movimientos recientemente.")
    else:
        print(f"{'FECHA':<17} {'PLACA':<12} {'MARCA':<15} {'ACCIÓN':<30}")
        print("-" * 74)
        
        for mov in movimientos:
            fecha_obj = datetime.strptime(mov['fecha'], "%Y-%m-%d %H:%M:%S")
            fecha_formateada = fecha_obj.strftime("%d/%m/%Y %H:%M")
            
            accion = mov.get('accion', 'N/A')
            if len(accion) > 28:
                accion = accion[:27] + "..."

            placa = mov.get('equipo_placa', 'N/A')
            marca = mov.get('marca', 'N/A') or 'N/A'

            print(f"{fecha_formateada:<17} {placa:<12} {marca:<15} {accion:<30}")

def menu_reportes_excel(usuario: str):
    """Muestra el menú para generar los reportes de inventario en Excel."""
    print("\n--- Generar Reportes en Excel ---")
    opciones = [
        "1. Reporte de Inventario Actual (Equipos Activos)",
        "2. Reporte de Equipos Devueltos a Proveedor",
        "3. Reporte Histórico Completo de Equipos (Log)",
        "4. Volver"
    ]
    print("\n".join(opciones))

@requiere_permiso("generar_reporte")
def generar_excel_inventario(usuario: str) -> None:
    """Genera un reporte Excel con los equipos activos."""
    try:
        inventario = db_manager.get_equipos_activos()
        if not inventario:
            print("\nNo hay equipos activos para generar un reporte.")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario de Equipos"

        # ... (Lógica de creación de Excel sin cambios) ...
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        encabezados = ["FECHA REGISTRO", "PLACA", "TIPO", "MARCA", "MODELO", "SERIAL", "ESTADO", "FECHA ÚLTIMO CAMBIO", "USUARIO ÚLTIMO CAMBIO", "ASIGNADO A", "EMAIL", "ÚLTIMA OBSERVACIÓN"]
        column_widths = {'A': 25, 'B': 15, 'C': 25, 'D': 25, 'E': 25, 'F': 30, 'G': 30, 'H': 25, 'I': 25, 'J': 30, 'K': 30, 'L': 80}
        for col, width in column_widths.items(): ws.column_dimensions[col].width = width
        for col_num, encabezado in enumerate(encabezados, 1):
            celda = ws[f"{get_column_letter(col_num)}1"]
            celda.value = encabezado
            celda.fill = header_fill; celda.font = header_font; celda.alignment = Alignment(horizontal='center'); celda.border = border
        colores_estado = {"Disponible": "C6EFCE", "Asignado": "FFEB9C", "En préstamo": "DDEBF7", "En mantenimiento": "FCE4D6", "Pendiente Devolución a Proveedor": "FFFFCC"}
        for row_num, equipo in enumerate(inventario, 2):
            ultimo_movimiento = db_manager.get_last_movimiento_by_placa(equipo['placa'])
            fecha_ult_cambio, usuario_ult_cambio = "N/A", "N/A"
            ultima_observacion = equipo.get('observaciones', 'N/A')
            if ultimo_movimiento:
                fecha_obj = datetime.strptime(ultimo_movimiento['fecha'], "%Y-%m-%d %H:%M:%S")
                fecha_ult_cambio = fecha_obj.strftime("%d/%m/%Y %H:%M")
                usuario_ult_cambio = ultimo_movimiento.get('usuario', 'N/A')
                ultima_observacion = ultimo_movimiento.get('detalles', ultima_observacion)
            data_row = [equipo.get('fecha_registro', 'N/A'), equipo.get('placa', 'N/A'), equipo.get('tipo', 'N/A'), equipo.get('marca', 'N/A'), equipo.get('modelo', 'N/A'), equipo.get('serial', 'N/A'), equipo.get('estado', 'N/A'), fecha_ult_cambio, usuario_ult_cambio, equipo.get('asignado_a', ''), equipo.get('email_asignado', ''), ultima_observacion]
            for col_num, cell_value in enumerate(data_row, 1):
                ws.cell(row=row_num, column=col_num, value=cell_value).border = border
            estado_celda = ws.cell(row=row_num, column=7)
            if color_hex := colores_estado.get(equipo.get('estado')):
                estado_celda.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        ws.freeze_panes = "A2"
        # ... (Fin de la lógica de Excel) ...

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            ruta_temporal = tmp.name

        registrar_movimiento_sistema("Reporte Inventario Activo", f"Generado reporte con {len(inventario)} equipos", usuario)
        
        # MEJORA: Imprimimos la ruta del archivo en lugar de intentar abrirlo.
        print(f"\n✅ Reporte de inventario activo generado.")
        print(f"Ruta en el servidor: {ruta_temporal}")

    except Exception as e:
        print(f"\n❌ Error al generar el reporte Excel: {str(e)}")

@requiere_permiso("generar_reporte")
def generar_excel_devueltos_proveedor(usuario: str) -> None:
    try:
        inventario_devuelto = db_manager.get_equipos_devueltos()
        if not inventario_devuelto:
            print("\nNo hay equipos devueltos al proveedor para reportar.")
            return
            
        wb = Workbook()
        # ... (Lógica de creación de Excel sin cambios) ...
        ws = wb.active
        ws.title = "Equipos Devueltos"
        header_fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
        header_font = Font(color="000000", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        encabezados = ["PLACA", "TIPO", "MARCA", "MODELO", "SERIAL", "FECHA DEVOLUCIÓN", "MOTIVO DEVOLUCIÓN", "ÚLTIMA OBSERVACIÓN"]
        column_widths = {'A': 15, 'B': 25, 'C': 25, 'D': 25, 'E': 30, 'F': 25, 'G': 25, 'H': 80}
        for col, width in column_widths.items(): ws.column_dimensions[col].width = width
        for col_num, encabezado in enumerate(encabezados, 1):
            celda = ws[f"{get_column_letter(col_num)}1"]
            celda.value = encabezado
            celda.fill = header_fill; celda.font = header_font; celda.alignment = Alignment(horizontal='center'); celda.border = border
        for row_num, equipo in enumerate(inventario_devuelto, 2):
            data_row = [equipo.get('placa', 'N/A'), equipo.get('tipo', 'N/A'), equipo.get('marca', 'N/A'), equipo.get('modelo', 'N/A'), equipo.get('serial', 'N/A'), equipo.get('fecha_devolucion_proveedor', 'N/A'), equipo.get('motivo_devolucion', 'N/A'), equipo.get('observaciones', 'N/A')]
            for col_num, cell_value in enumerate(data_row, 1):
                ws.cell(row=row_num, column=col_num, value=cell_value).border = border
        ws.freeze_panes = "A2"
        # ... (Fin de la lógica de Excel) ...
        
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            ruta_temporal = tmp.name
        
        registrar_movimiento_sistema("Reporte Equipos Devueltos", f"Generado reporte con {len(inventario_devuelto)} equipos devueltos", usuario)

        # MEJORA: Imprimimos la ruta del archivo.
        print(f"\n✅ Reporte de equipos devueltos generado.")
        print(f"Ruta en el servidor: {ruta_temporal}")

    except Exception as e:
        print(f"\n❌ Error al generar el reporte de equipos devueltos: {str(e)}")

@requiere_permiso("ver_historico")
def generar_excel_historico(usuario: str):
    try:
        log_equipos = db_manager.get_all_log_inventario()
        if not log_equipos:
            print("\nNo hay movimientos de equipos para exportar.")
            return

        wb = Workbook()
        # ... (Lógica de creación de Excel sin cambios) ...
        ws = wb.active
        ws.title = "Histórico de Movimientos"
        header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        encabezados = ["FECHA", "PLACA EQUIPO", "ACCIÓN", "USUARIO", "DETALLES"]
        for col_num, encabezado in enumerate(encabezados, 1):
            celda = ws[f"{get_column_letter(col_num)}1"]
            celda.value = encabezado
            celda.fill = header_fill; celda.font = header_font; celda.alignment = Alignment(horizontal='center'); celda.border = border
        ws.column_dimensions['A'].width = 25; ws.column_dimensions['B'].width = 20; ws.column_dimensions['C'].width = 25; ws.column_dimensions['D'].width = 20; ws.column_dimensions['E'].width = 80
        for row_num, mov in enumerate(log_equipos, 2):
            fecha_obj = datetime.strptime(mov['fecha'], "%Y-%m-%d %H:%M:%S")
            fecha_formateada = fecha_obj.strftime("%d/%m/%Y %H:%M")
            ws.cell(row=row_num, column=1, value=fecha_formateada).border = border
            ws.cell(row=row_num, column=2, value=mov.get('equipo_placa', 'N/A')).border = border
            ws.cell(row=row_num, column=3, value=mov.get('accion', 'N/A')).border = border
            ws.cell(row=row_num, column=4, value=mov.get('usuario', 'N/A')).border = border
            ws.cell(row=row_num, column=5, value=mov.get('detalles', '')).border = border
        ws.freeze_panes = "A2"
        # ... (Fin de la lógica de Excel) ...
        
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            ruta_temporal = tmp.name

        registrar_movimiento_sistema("Reporte Histórico Equipos", "Generado reporte de histórico de equipos", usuario)

        # MEJORA: Imprimimos la ruta del archivo.
        print(f"\n✅ Reporte histórico de equipos generado.")
        print(f"Ruta en el servidor: {ruta_temporal}")

    except Exception as e:
        print(f"\n❌ Error al generar el reporte de histórico: {str(e)}")


@requiere_permiso("ver_inventario")
def ver_inventario_consola():
    """Muestra la primera página del inventario activo."""
    print("\n--- Inventario Actual de Equipos Activos (Página 1) ---")
    
    total_equipos = db_manager.count_equipos_activos()
    if total_equipos == 0:
        print("\nEl inventario activo está vacío.")
        return
        
    # Mostramos solo la primera página para esta versión
    inventario = db_manager.get_equipos_activos_paginated(page=1, page_size=20)

    print(f"{'PLACA':<15} {'TIPO':<20} {'ESTADO':<35} {'ASIGNADO A'}")
    print("=" * 90)
    
    for equipo in inventario:
        estado = equipo['estado']
        asignado_a = equipo.get('asignado_a') or 'N/A'
        print(f"{equipo['placa']:<15} {equipo['tipo']:<20} {estado:<35} {asignado_a}")

    print(f"\nMostrando {len(inventario)} de {total_equipos} equipos.")
    # La paginación interactiva requeriría una máquina de estados más compleja en app.py

def generar_excel_historico_equipo(usuario: str, equipo: Equipo):
    """Genera y reporta la ruta de un Excel para el historial de un solo equipo."""
    try:
        log_equipo = db_manager.get_log_by_placa(equipo.placa)
        if not log_equipo:
            print(f"\nNo hay historial para el equipo {equipo.placa}.")
            return

        wb = Workbook()
        # ... (Lógica de creación de Excel sin cambios) ...
        ws = wb.active
        ws.title = f"Historial {equipo.placa}"
        header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        encabezados = ["FECHA", "ACCIÓN", "USUARIO", "DETALLES"]
        for col_num, encabezado in enumerate(encabezados, 1):
            celda = ws[f"{get_column_letter(col_num)}1"]
            celda.value = encabezado
            celda.fill = header_fill; celda.font = header_font; celda.alignment = Alignment(horizontal='center'); celda.border = border
        ws.column_dimensions['A'].width = 25; ws.column_dimensions['B'].width = 25; ws.column_dimensions['C'].width = 20; ws.column_dimensions['D'].width = 80
        for row_num, mov in enumerate(log_equipo, 2):
            fecha_obj = datetime.strptime(mov['fecha'], "%Y-%m-%d %H:%M:%S")
            fecha_formateada = fecha_obj.strftime("%d/%m/%Y %H:%M")
            ws.cell(row=row_num, column=1, value=fecha_formateada).border = border
            ws.cell(row=row_num, column=2, value=mov.get('accion', 'N/A')).border = border
            ws.cell(row=row_num, column=3, value=mov.get('usuario', 'N/A')).border = border
            ws.cell(row=row_num, column=4, value=mov.get('detalles', '')).border = border
        ws.freeze_panes = "A2"
        # ... (Fin de la lógica de Excel) ...
        
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            ruta_temporal = tmp.name

        registrar_movimiento_sistema("Reporte Histórico Individual", f"Generado reporte para placa {equipo.placa}", usuario)
        
        # MEJORA: Imprimimos la ruta del archivo.
        print(f"\n✅ Reporte de historial para {equipo.placa} generado.")
        print(f"Ruta en el servidor: {ruta_temporal}")

    except Exception as e:
        print(f"\n❌ Error al generar el historial del equipo: {str(e)}")