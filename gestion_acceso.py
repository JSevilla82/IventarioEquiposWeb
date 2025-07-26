# gestion_acceso.py
import bcrypt
import re
import tempfile
import webbrowser
from functools import wraps
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

# Usaremos la sesión de Flask para obtener información del usuario
from flask import session

from database import db_manager, Usuario, registrar_movimiento_sistema

# --- CONTROL DE ACCESO BASADO EN ROLES (RBAC) ---
ROLES_PERMISOS = {
    "Administrador": {
        "registrar_equipo", "ver_inventario", "gestionar_equipo", "ver_historico",
        "generar_reporte", "gestionar_usuarios", "eliminar_equipo",
        "devolver_a_proveedor", "aprobar_devoluciones", "gestionar_pendientes",
        "configurar_sistema"
    },
    "Gestor": {
        "registrar_equipo", "ver_inventario", "gestionar_equipo", "ver_historico",
        "generar_reporte", "devolver_a_proveedor"
    },
    "Visualizador": {
        "ver_inventario", "ver_historico", "generar_reporte"
    }
}

def requiere_permiso(permiso: str):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            # MEJORA: Obtenemos el rol desde la sesión del navegador, no de una variable global.
            rol_usuario = session.get('rol')
            if not rol_usuario:
                print("\n❌ Acceso denegado. No hay un rol de usuario definido en la sesión.")
                return

            if permiso in ROLES_PERMISOS.get(rol_usuario, {}):
                return func(*args, **kwargs)
            else:
                print(f"\n❌ Permiso denegado. Su rol '{rol_usuario}' no tiene el permiso '{permiso}'.")
                return
        return wrapper
    return decorator

# --- FUNCIONES DE AUTENTICACIÓN Y GESTIÓN DE USUARIOS ---

def hash_contrasena(contrasena: str) -> str:
    return bcrypt.hashpw(contrasena.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verificar_contrasena(contrasena: str, hash_almacenado: str) -> bool:
    return bcrypt.checkpw(contrasena.encode('utf-8'), hash_almacenado.encode('utf-8'))

def validar_contrasena(contrasena: str) -> bool:
    if len(contrasena) < 8: return False
    if not re.search(r'[A-Za-z]', contrasena): return False
    if not re.search(r'[0-9]', contrasena): return False
    return True

def login(nombre_usuario, contrasena):
    """
    Función de login adaptada para el servidor web.
    No contiene bucles ni inputs, solo valida las credenciales una vez.
    Retorna el nombre de usuario si es exitoso, o None si falla.
    """
    if not nombre_usuario or not contrasena:
        return None

    user_data = db_manager.get_user_by_username(nombre_usuario)
    
    if user_data and verificar_contrasena(contrasena, user_data['contrasena_hash']):
        if not user_data['is_active']:
            print("❌ Su cuenta de usuario está bloqueada. Contacte a un administrador.")
            return None
        
        # El manejo del cambio de contraseña requerido se haría en app.py después del login
        return nombre_usuario
    
    return None

def inicializar_admin_si_no_existe():
    if not db_manager.get_user_by_username("admin"):
        print("\nCreando usuario administrador inicial 'admin'...")
        admin_pass_hash = hash_contrasena("adminpass")
        admin_user = Usuario("admin", admin_pass_hash, "Administrador", "Administrador Principal", True, True)
        db_manager.insert_user(admin_user)
        print("✅ Usuario 'admin' creado con contraseña 'adminpass'. Por favor, cámbiela.")

# --- MENÚS (ADAPTADOS PARA IMPRIMIR Y TERMINAR) ---
# Estas funciones ahora solo muestran su contenido. La lógica para manejar la
# opción seleccionada por el usuario se gestionará en app.py en el futuro.

@requiere_permiso("gestionar_usuarios")
def menu_usuarios(usuario_actual: str):
    """Muestra la lista de usuarios y las opciones disponibles."""
    print("\n--- Gestión de Usuarios ---")
    usuarios = db_manager.get_all_users()
    
    print(f"{'USUARIO':<20} {'NOMBRE COMPLETO':<30} {'ESTADO'}")
    print("-" * 65)
    if not usuarios:
        print("No hay usuarios registrados.")
    else:
        for user in usuarios:
            estado = "Activo" if user['is_active'] else "Bloqueado"
            nombre_completo = user.get('nombre_completo') or 'N/A'
            print(f"{user['nombre_usuario']:<20} {nombre_completo:<30} {estado}")
    print("-" * 65)
    
    opciones = [
        "1. Registrar nuevo usuario",
        "2. Gestionar un usuario existente",
        "3. Volver"
    ]
    print("\nOpciones de Gestión de Usuarios:")
    print("\n".join(opciones))
    # En un futuro, aquí se cambiaría el estado en app.py para esperar la opción.

@requiere_permiso("ver_historico")
def menu_ver_log_sistema(usuario: str):
    print("\n--- Log de Actividad del Sistema ---")
    # Para la versión web, la generación de Excel es la única opción viable.
    generar_excel_log_sistema(usuario)

@requiere_permiso("ver_historico")
def generar_excel_log_sistema(usuario: str):
    try:
        log_sistema = db_manager.get_all_log_sistema()

        if not log_sistema:
            print("\nNo hay actividad del sistema para exportar.")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Log del Sistema"

        # ... (lógica de creación de Excel sin cambios) ...
        header_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
        header_font = Font(color="000000", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        encabezados = ["FECHA", "ACCIÓN", "USUARIO", "DETALLES"]
        for col_num, encabezado in enumerate(encabezados, 1):
            col_letra = get_column_letter(col_num)
            celda = ws[f"{col_letra}1"]
            celda.value = encabezado
            celda.fill = header_fill; celda.font = header_font; celda.alignment = Alignment(horizontal='center'); celda.border = border
        ws.column_dimensions['A'].width = 25; ws.column_dimensions['B'].width = 25; ws.column_dimensions['C'].width = 20; ws.column_dimensions['D'].width = 80
        for row_num, mov in enumerate(log_sistema, 2):
            fecha_obj = datetime.strptime(mov['fecha'], "%Y-%m-%d %H:%M:%S")
            fecha_formateada = fecha_obj.strftime("%d/%m/%Y %H:%M")
            ws.cell(row=row_num, column=1, value=fecha_formateada).border = border
            ws.cell(row=row_num, column=2, value=mov.get('accion', 'N/A')).border = border
            ws.cell(row=row_num, column=3, value=mov.get('usuario', 'N/A')).border = border
            ws.cell(row=row_num, column=4, value=mov.get('detalles', '')).border = border
        ws.freeze_panes = "A2"
        # ... (fin de la lógica de creación de Excel) ...

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            ruta_temporal = tmp.name

        # MEJORA: En lugar de intentar abrir el archivo, imprimimos su ruta.
        print(f"\n✅ Reporte de Log del Sistema generado exitosamente.")
        print(f"El archivo se encuentra en la siguiente ruta del servidor: {ruta_temporal}")

    except Exception as e:
        print(f"\n❌ Error al generar el log del sistema: {str(e)}")


@requiere_permiso("configurar_sistema")
def menu_configuracion_sistema(usuario: str):
    """Muestra el menú de configuración del sistema."""
    print("\n--- Configuración del Sistema ---")
    opciones = [
        "1. Gestionar Tipos de Equipo",
        "2. Gestionar Marcas",
        "3. Gestionar Dominios de Correo",
        "4. Volver"
    ]
    print("\n".join(opciones))
    # Aquí iría la lógica para manejar la opción del usuario en app.py

# Las funciones más complejas como registrar_usuario, cambiar_contrasena_usuario,
# gestionar_parametros, etc., requerirán una adaptación más profunda con la
# máquina de estados de app.py para manejar la entrada de múltiples campos.
# Por ahora, las dejamos fuera para mantener la simplicidad de este paso.
def cambiar_contrasena_usuario(usuario: str):
    print("\n--- Cambio de Contraseña ---")
    print("Esta función requiere un flujo interactivo que se implementará próximamente.")

def menu_gestion_acceso_sistema(usuario: str):
    """Menú principal para acceso y sistema (adaptado)."""
    print("\n--- Módulo de Acceso y Sistema ---")
    rol_actual = session.get('rol', 'Visualizador')
    
    opciones_disponibles = []
    if "gestionar_usuarios" in ROLES_PERMISOS[rol_actual]: opciones_disponibles.append("1. Gestión de usuarios")
    if "configurar_sistema" in ROLES_PERMISOS[rol_actual]: opciones_disponibles.append("2. Configuración del Sistema")
    if "ver_historico" in ROLES_PERMISOS[rol_actual]: opciones_disponibles.append("3. Ver Log de Actividad del Sistema")
    opciones_disponibles.append("4. Cambiar mi contraseña")
    opciones_disponibles.append("5. Volver al menú principal")
    
    print("\n".join(opciones_disponibles))
    # La lógica para manejar estas opciones se añadiría a app.py